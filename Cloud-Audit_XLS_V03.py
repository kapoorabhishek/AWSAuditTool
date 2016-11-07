#!/usr/bin/python

import boto.ec2
import boto.ec2.elb
import boto.rds
from boto.s3.connection import S3Connection
from openpyxl import load_workbook
import datetime

# Function to get EC2 instancess
def get_ec2_instances(wb_ref):

  #conn = boto.ec2.connect_to_region("eu-west-1",aws_access_key_id="AKIAIJQBPWWWCHCZHL4A",aws_secret_access_key="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")
  conn = boto.ec2.connect_to_region("eu-west-1")
  reservations = conn.get_all_reservations()

  sheet1 = wb_ref.create_sheet("EC2")

  #Add Headers
  sheet1.cell(row=1,column=1,value="NAME")
  sheet1.cell(row=1,column=2,value="INSTANCE ID")
  sheet1.cell(row=1,column=3,value="INSTANCE TYPE")
  sheet1.cell(row=1,column=4,value="INSTANCE STATE")
  sheet1.cell(row=1,column=5,value="PUBLIC IP")
  sheet1.cell(row=1,column=6,value="STACK NAME")
  sheet1.cell(row=1,column=7,value="$$ PER MONTH")

  i = 2
  for reservation in reservations:
    for instance in reservation.instances:

      if 'Name' in instance.tags:
        sheet1.cell(row=i, column=1, value=instance.tags['Name'] )

      sheet1.cell(row=i,column=2,value=instance.id)
      sheet1.cell(row=i,column=3,value=instance.instance_type)
      sheet1.cell(row=i,column=4,value=instance.state)
      sheet1.cell(row=i,column=5,value=instance.ip_address)


      if 'aws:cloudformation:stack-name' in instance.tags:
        sheet1.cell(row=i,column=6,value=instance.tags['aws:cloudformation:stack-name'])
      elif 'opsworks:stack' in instance.tags:
        sheet1.cell(row=i,column=6,value=instance.tags['opsworks:stack'])

      sheet1.cell(row=i,column=7,value=("=24*30*vlookup(C" + str(i) + ",EC2Pricing!A$2:F$51,6,FALSE)"))
      i=i+1


# Get RDS List
def get_rds_instances(wb_ref):
    rdsconn = boto.rds.connect_to_region("eu-west-1")
    instances = rdsconn.get_all_dbinstances()

    sheet2 = wb_ref.create_sheet("RDS")

    #Add Headers
    sheet2.cell(row=1,column=1,value="NAME")
    sheet2.cell(row=1,column=2,value="ENGINE")
    sheet2.cell(row=1,column=3,value="CLASS")
    sheet2.cell(row=1,column=4,value="LAUNCH_TIME")
    sheet2.cell(row=1,column=5,value="ALLOCATED STORAGE")
    sheet2.cell(row=1,column=6,value="MULTI AZ")
    sheet2.cell(row=1,column=7,value="PUBLICLY ACCESSIBLE")
    sheet2.cell(row=1,column=8,value="$$ PER MONTH")

    i=2
    for instance in instances:
        sheet2.cell(row=i,column=1,value=instance.id)
        sheet2.cell(row=i,column=2,value=instance.engine)
        sheet2.cell(row=i,column=3,value=instance.instance_class)
        sheet2.cell(row=i,column=4,value=instance.create_time)
        sheet2.cell(row=i,column=5,value=instance.allocated_storage)
        sheet2.cell(row=i,column=6,value=instance.multi_az)
        sheet2.cell(row=i,column=7,value=instance.PubliclyAccessible)
        sheet2.cell(row=i,column=8,value=("=24*30*vlookup(B" + str(i) + "&C" + str(i) + "&F" + str(i) + ",RDSPricing!A$2:E$85,5,FALSE)" ))
        i=i+1


def get_elb_instances(wb_ref):
    elb = boto.ec2.elb.connect_to_region('eu-west-1')
    balancers = elb.get_all_load_balancers()

    sheet3 = wb_ref.create_sheet("ELB")

    #Add Headers
    sheet3.cell(row=1,column=1,value="NAME")
    sheet3.cell(row=1,column=2,value="INSTANCES")
    sheet3.cell(row=1,column=3,value="SCHEME")
    sheet3.cell(row=1,column=4,value="CREATED_TIME")
    sheet3.cell(row=1,column=5,value="$$ PER MONTH")

    i=2
    for balancer in balancers:
        sheet3.cell(row=i,column=1,value=balancer.name)
        inst = ''
        for instance in balancer.instances:
            inst = instance.id + ',' + inst
        sheet3.cell(row=i,column=2,value=inst)
        sheet3.cell(row=i,column=3,value=balancer.scheme)
        sheet3.cell(row=i,column=4,value=balancer.created_time)
        sheet3.cell(row=i,column=5,value="=0.028*24*30")
        i=i+1


def get_ec2_volumes(wb_ref):
    conn = boto.ec2.connect_to_region('eu-west-1')
    volumes = conn.get_all_volumes()

    sheet4 = wb_ref.create_sheet("Volumes")

    #Add Headers
    sheet4.cell(row=1,column=1,value="NAME")
    sheet4.cell(row=1,column=2,value="SIZE")
    sheet4.cell(row=1,column=3,value="ENCRYPTED")
    sheet4.cell(row=1,column=4,value="VOLUME TYPE")
    sheet4.cell(row=1,column=5,value="IOPS")
    sheet4.cell(row=1,column=6,value="CREATED_TIME")
    sheet4.cell(row=1,column=7,value="SNAPSHOT ID")
    sheet4.cell(row=1,column=8,value="INSTANCE ATTACHED")
    sheet4.cell(row=1,column=9,value="TAG: NAME")
    sheet4.cell(row=1,column=10,value="$$ PER MONTH")

    i=2
    for volume in volumes:
        sheet4.cell(row=i,column=1,value=volume.id)
        sheet4.cell(row=i,column=2,value=volume.size)
        sheet4.cell(row=i,column=3,value=volume.encrypted)
        sheet4.cell(row=i,column=4,value=volume.type)
        sheet4.cell(row=i,column=5,value=volume.iops)
        sheet4.cell(row=i,column=6,value=volume.create_time)
        sheet4.cell(row=i,column=7,value=volume.snapshot_id)
        inst=""
        if volume.attachment_state() == u'attached':
            inst=volume.attach_data.instance_id
        sheet4.cell(row=i,column=8,value=inst)
        if 'Name' in volume.tags:
          sheet4.cell(row=i,column=9, value=volume.tags['Name'] )
        formula = '=if(D' + str(i) + '="iO1",(B' + str(i) + '*vlookup(D' + str(i) + ',VolumePricing!A$2:B$5,2,FALSE))+(E' + str(i) + '*vlookup("IOPS",VolumePricing!A$2:B$5,2,FALSE)),B' + str(i) + '*vlookup(D' + str(i) + ',VolumePricing!A$2:B$5,2,FALSE))'

        sheet4.cell(row=i,column=10, value=formula)
        i=i+1


def get_ec2_snapshots(wb_ref):
    conn = boto.ec2.connect_to_region("eu-west-1")
    snaps = conn.get_all_snapshots()

    sheet5 = wb_ref.create_sheet("Snapshots")

    #Add Headers
    sheet5.cell(row=1,column=1,value="SNAPSHOT ID")
    sheet5.cell(row=1,column=2,value="SIZE")
    sheet5.cell(row=1,column=3,value="ENCRYPTED")
    sheet5.cell(row=1,column=4,value="DESCRIPTION")
    sheet5.cell(row=1,column=5,value="START_TIME")
    sheet5.cell(row=1,column=6,value="VOLUME ID")
    sheet5.cell(row=1,column=7,value="STATUS")

    i=2
    for snap in snaps:
        sheet5.cell(row=i,column=1,value=snap.id)
        sheet5.cell(row=i,column=2,value=snap.volume_size)
        sheet5.cell(row=i,column=3,value=snap.encrypted)
        sheet5.cell(row=i,column=4,value=snap.description)
        sheet5.cell(row=i,column=5,value=snap.start_time)
        sheet5.cell(row=i,column=6,value=snap.volume_id)
        sheet5.cell(row=i,column=7,value=snap.status)
        i=i+1


def get_s3(wb_ref):
    conn = S3Connection()
    buckets = conn.get_all_buckets()

    sheet6 = wb_ref.create_sheet("S3")

    #Add Headers
    sheet6.cell(row=1,column=1,value="BUCKET NAME")
    sheet6.cell(row=1,column=2,value="CREATION DATE")

    i=2
    for bucket in buckets:
        sheet6.cell(row=i,column=1,value=bucket.name)
        sheet6.cell(row=i,column=2,value=bucket.creation_date)
        #b = conn.get_bucket(bucket.name)
        #size=0
        #for key in b.list():
        #    size += key.size
        #sheet6.write(i,2,size)
        i=i+1


#Main

## Parse the command line options
#parser = optparse.OptionParser()
#parser.add_option('-a', dest="AWSAccount", help="The AWS Account to be audited")
#(opts, args) = parser.parse_args()
#if not (opts.AWSAccount):
    #print "Missing Argument - Account Name."
    #print "Usage:"
    #print "\tpython Cloud-Audit_XLS.py -a <Account Name>"
    #exit()
#os.system("export AWS_DEFAULT_PROFILE=opts.AWSAccount")

wb=load_workbook("AuditTemplate.xlsx")
get_ec2_instances(wb)
get_ec2_volumes(wb)
get_rds_instances(wb)
get_elb_instances(wb)
get_s3(wb)
get_ec2_snapshots(wb)

name = "Audit_" + datetime.datetime.now().strftime('%Y-%m-%d-%H%M') + ".xlsx"
wb.save(name)
